import os
import sqlite3
from openpyxl import load_workbook
import unicodedata
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, 'instance', 'app.db')
PHAN_GIAO_FILE = 'phan_giao.xlsx'
FORM_FILE = 'forms.xlsx'

def build_sheet_mapping():
    wb = load_workbook(FORM_FILE, data_only=True)
    mapping = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('BM'):
            code_part = sheet_name[2:]
            if code_part.isdigit():
                num = int(code_part)
                mapping[f'BM.P4.15.{num:02d}'] = sheet_name
            elif '_' in code_part:
                base, pha = code_part.split('_')
                num = int(base)
                mapping[f'BM.P4.15.{num:02d}_{pha}'] = sheet_name
    wb.close()
    return mapping

def get_or_create_user(cursor, fullname, role):
    nfkd = unicodedata.normalize('NFKD', fullname)
    ascii_name = nfkd.encode('ASCII', 'ignore').decode('utf-8')
    username = re.sub(r'\s+', '', ascii_name).lower()
    cursor.execute("SELECT id FROM users WHERE username = ?", (username,))
    user = cursor.fetchone()
    if user:
        return user[0]
    else:
        cursor.execute("INSERT INTO users (username, password, fullname, role) VALUES (?, ?, ?, ?)", (username, '123', fullname, role))
        return cursor.lastrowid

def init_db():
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            fullname TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('danh_gia', 'tham_tra'))
        );
        CREATE TABLE IF NOT EXISTS assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            role TEXT NOT NULL,
            FOREIGN KEY (user_id) REFERENCES users (id)
        );
        CREATE TABLE IF NOT EXISTS evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            col_letter TEXT NOT NULL,
            value TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        );
        CREATE TABLE IF NOT EXISTS review_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reviewer_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            comment TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (reviewer_id) REFERENCES users (id)
        );
        CREATE TABLE IF NOT EXISTS suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            sheet_name TEXT NOT NULL,
            suggestion TEXT,
            reviewer_comment TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    conn.commit()

    sheet_mapping = build_sheet_mapping()

    wb = load_workbook(PHAN_GIAO_FILE, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=7, values_only=True):
        stt = row[0]
        if stt is None or (isinstance(stt, (int, float)) and stt == ''):
            break
        ma_bieu_mau = row[2]
        ten_danh_gia = row[4]
        ten_tham_tra = row[5]

        if not ma_bieu_mau:
            continue

        if ma_bieu_mau in ['BM.P4.15.18', 'BM.P4.15.19']:
            base_num = ma_bieu_mau.split('.')[-1]
            sheet_names = [f'BM{base_num}_a', f'BM{base_num}_b', f'BM{base_num}_c']
        else:
            sheet_name = sheet_mapping.get(ma_bieu_mau)
            if not sheet_name:
                print(f"Không tìm thấy sheet cho mã {ma_bieu_mau}")
                continue
            sheet_names = [sheet_name]

        if ten_danh_gia:
            user_id = get_or_create_user(cursor, ten_danh_gia, 'danh_gia')
            for sname in sheet_names:
                cursor.execute("SELECT id FROM assignments WHERE user_id = ? AND sheet_name = ? AND role = ?", (user_id, sname, 'danh_gia'))
                if not cursor.fetchone():
                    cursor.execute("INSERT INTO assignments (user_id, sheet_name, role) VALUES (?, ?, ?)", (user_id, sname, 'danh_gia'))
            conn.commit()

        if ten_tham_tra:
            user_id = get_or_create_user(cursor, ten_tham_tra, 'tham_tra')
            for sname in sheet_names:
                cursor.execute("SELECT id FROM assignments WHERE user_id = ? AND sheet_name = ? AND role = ?", (user_id, sname, 'tham_tra'))
                if not cursor.fetchone():
                    cursor.execute("INSERT INTO assignments (user_id, sheet_name, role) VALUES (?, ?, ?)", (user_id, sname, 'tham_tra'))
            conn.commit()

    wb.close()
    conn.close()
    print("Khởi tạo database thành công!")

if __name__ == '__main__':
    init_db()
