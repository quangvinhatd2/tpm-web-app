import os
import re
import json
from io import BytesIO
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from unicodedata import normalize
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2.pool import SimpleConnectionPool
from contextlib import contextmanager
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = 'tpm-secure-key-2024'

DATABASE_URL = os.environ.get('DATABASE_URL')
if not DATABASE_URL:
    raise Exception("DATABASE_URL environment variable not set")

FORMS_FILE = 'forms.xlsx'
PHAN_GIAO_FILE = 'phan_giao.xlsx'

# Tạo connection pool
db_pool = SimpleConnectionPool(1, 10, DATABASE_URL, sslmode='require')

@contextmanager
def get_db_connection():
    conn = db_pool.getconn()
    try:
        # Kiểm tra kết nối còn sống, nếu không thì thay thế
        if conn.closed:
            db_pool.putconn(conn)
            conn = db_pool.getconn()
        # Đặt cursor_factory
        conn.cursor_factory = RealDictCursor
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        db_pool.putconn(conn)

# -------------------- Hàm xử lý Excel an toàn --------------------
def safe_load_workbook(filepath, read_only=False):
    """Load workbook với xử lý lỗi và read_only cho file lớn"""
    try:
        if read_only:
            return load_workbook(filepath, read_only=True, data_only=True)
        else:
            return load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Lỗi đọc file {filepath}: {e}")
        return None

def build_sheet_mapping():
    wb = safe_load_workbook(FORMS_FILE, read_only=True)
    if not wb:
        return {}
    mapping = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('BM'):
            code_part = sheet_name[2:]
            if code_part.isdigit():
                num = int(code_part)
                mapping[f'BM.P4.15.{num:02d}'] = sheet_name
            elif '_' in code_part:
                base, pha = code_part.split('_')
                num = int(base)
                mapping[f'BM.P4.15.{num:02d}_{pha}'] = sheet_name
    wb.close()
    return mapping

def build_reverse_mapping():
    wb = safe_load_workbook(FORMS_FILE, read_only=True)
    if not wb:
        return {}
    rev_map = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('BM'):
            code_part = sheet_name[2:]
            if code_part.isdigit():
                num = int(code_part)
                rev_map[sheet_name] = f'BM.P4.15.{num:02d}'
            elif '_' in code_part:
                base, pha = code_part.split('_')
                num = int(base)
                rev_map[sheet_name] = f'BM.P4.15.{num:02d}_{pha}'
    wb.close()
    return rev_map

def get_sheet_data(sheet_name):
    try:
        wb = safe_load_workbook(FORMS_FILE, read_only=True)
        if not wb or sheet_name not in wb.sheetnames:
            return None, None, None
        ws = wb[sheet_name]
        headers = [{col: ws[f'{col}{r}'].value for col in 'ABCDEF'} for r in range(1, 8)]
        rows = []
        extra = []
        # Giới hạn số dòng đọc để tránh treo (tối đa 500 dòng)
        max_row = min(ws.max_row, 500)
        for r_idx in range(10, max_row + 1):
            row_data = {col: ws[f'{col}{r_idx}'].value or '' for col in 'ABCDEF'}
            if not any(str(v).strip() for v in row_data.values()):
                # Phần extra bỏ qua để tránh lỗi
                break
            rows.append(row_data)
        wb.close()
        return headers, rows, extra
    except Exception as e:
        print(f"Lỗi đọc sheet {sheet_name}: {e}")
        return None, None, None

# -------------------- Các hàm xử lý database --------------------
def unsigned_user(text):
    if not text:
        return None
    text = text.strip()
    replacements = {
        'Đ': 'D', 'đ': 'd',
        'À': 'A', 'Á': 'A', 'Â': 'A', 'Ã': 'A',
        'à': 'a', 'á': 'a', 'â': 'a', 'ã': 'a',
        'È': 'E', 'É': 'E', 'Ê': 'E',
        'è': 'e', 'é': 'e', 'ê': 'e',
        'Ì': 'I', 'Í': 'I', 'ì': 'i', 'í': 'i',
        'Ò': 'O', 'Ó': 'O', 'Ô': 'O', 'Õ': 'O',
        'ò': 'o', 'ó': 'o', 'ô': 'o', 'õ': 'o',
        'Ù': 'U', 'Ú': 'U', 'ù': 'u', 'ú': 'u',
        'Ý': 'Y', 'ý': 'y',
        'Ă': 'A', 'ă': 'a', 'Ắ': 'A', 'ắ': 'a', 'Ặ': 'A', 'ặ': 'a',
        'Ằ': 'A', 'ằ': 'a', 'Ẳ': 'A', 'ẳ': 'a', 'Ẵ': 'A', 'ẵ': 'a',
        'Ấ': 'A', 'ấ': 'a', 'Ầ': 'A', 'ầ': 'a', 'Ẩ': 'A', 'ẩ': 'a',
        'Ẫ': 'A', 'ẫ': 'a', 'Ậ': 'A', 'ậ': 'a',
        'Ơ': 'O', 'ơ': 'o', 'Ớ': 'O', 'ớ': 'o', 'Ờ': 'O', 'ờ': 'o',
        'Ở': 'O', 'ở': 'o', 'Ỡ': 'O', 'ỡ': 'o', 'Ợ': 'O', 'ợ': 'o',
        'Ố': 'O', 'ố': 'o', 'Ồ': 'O', 'ồ': 'o', 'Ổ': 'O', 'ổ': 'o',
        'Ỗ': 'O', 'ỗ': 'o', 'Ộ': 'O', 'ộ': 'o',
        'Ư': 'U', 'ư': 'u', 'Ứ': 'U', 'ứ': 'u', 'Ừ': 'U', 'ừ': 'u',
        'Ử': 'U', 'ử': 'u', 'Ữ': 'U', 'ữ': 'u', 'Ự': 'U', 'ự': 'u',
        'Ế': 'E', 'ế': 'e', 'Ề': 'E', 'ề': 'e', 'Ể': 'E', 'ể': 'e',
        'Ễ': 'E', 'ễ': 'e', 'Ệ': 'E', 'ệ': 'e',
        'Ỉ': 'I', 'ỉ': 'i', 'Ị': 'I', 'ị': 'i',
        'Ỳ': 'Y', 'ỳ': 'y', 'Ỷ': 'Y', 'ỷ': 'y',
        'Ỹ': 'Y', 'ỹ': 'y', 'Ỵ': 'Y', 'ỵ': 'y',
        'Ả': 'A', 'ả': 'a', 'Ạ': 'A', 'ạ': 'a',
        'Ẻ': 'E', 'ẻ': 'e', 'Ẽ': 'E', 'ẽ': 'e', 'Ẹ': 'E', 'ẹ': 'e',
        'Ỏ': 'O', 'ỏ': 'o', 'Ọ': 'O', 'ọ': 'o',
        'Ủ': 'U', 'ủ': 'u', 'Ụ': 'U', 'ụ': 'u',
    }
    result = ''
    for ch in text:
        if ch in replacements:
            result += replacements[ch]
        else:
            result += normalize('NFKD', ch).encode('ascii', 'ignore').decode('ascii')
    return re.sub(r'[^a-zA-Z0-9]', '', result).lower()

def init_db():
    if not os.path.exists(PHAN_GIAO_FILE):
        print("Không tìm thấy file phan_giao.xlsx")
        return
    mapping = build_sheet_mapping()
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DROP TABLE IF EXISTS users CASCADE")
        cur.execute("DROP TABLE IF EXISTS assignments CASCADE")
        cur.execute("""
            CREATE TABLE users (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE,
                password TEXT,
                fullname TEXT,
                role TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE assignments (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                sheet_name TEXT,
                role TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS evaluations (
                user_id INTEGER,
                sheet_name TEXT,
                row_index INTEGER,
                col_letter TEXT,
                value TEXT,
                PRIMARY KEY (user_id, sheet_name, row_index, col_letter)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS review_comments (
                reviewer_id INTEGER,
                sheet_name TEXT,
                row_index INTEGER,
                comment TEXT,
                PRIMARY KEY (reviewer_id, sheet_name, row_index)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS suggestions (
                sheet_name TEXT PRIMARY KEY,
                suggestion TEXT,
                reviewer_comment TEXT,
                reviewer_signature TEXT,
                checker_signature TEXT,
                locked_danh_gia INTEGER DEFAULT 0,
                locked_tham_tra INTEGER DEFAULT 0
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS archives (
                id SERIAL PRIMARY KEY,
                archive_date TEXT NOT NULL,
                table_name TEXT NOT NULL,
                row_data TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS history (
                id SERIAL PRIMARY KEY,
                sheet_name TEXT NOT NULL,
                role TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                saved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                snapshot TEXT NOT NULL
            )
        """)
        conn.commit()

        wb = safe_load_workbook(PHAN_GIAO_FILE)
        if not wb:
            print("Không thể đọc file phan_giao.xlsx")
            return
        ws = wb.active
        for row in range(8, ws.max_row + 1):
            ma_bieu_mau = str(ws[f'C{row}'].value or '').strip()
            name_eval = str(ws[f'E{row}'].value or '').strip()
            name_check = str(ws[f'F{row}'].value or '').strip()
            if not ma_bieu_mau or (not name_eval and not name_check):
                continue
            if ma_bieu_mau in ['BM.P4.15.18', 'BM.P4.15.19']:
                base_num = ma_bieu_mau.split('.')[-1]
                sheet_names = [f'BM{base_num}_a', f'BM{base_num}_b', f'BM{base_num}_c']
            else:
                sheet_name = mapping.get(ma_bieu_mau)
                if not sheet_name:
                    continue
                sheet_names = [sheet_name]
            if name_eval:
                uid = create_or_get_user(conn, name_eval, 'danh_gia')
                for sname in sheet_names:
                    cur.execute("INSERT INTO assignments (user_id, sheet_name, role) VALUES (%s, %s, %s)", (uid, sname, 'danh_gia'))
            if name_check:
                uid = create_or_get_user(conn, name_check, 'tham_tra')
                for sname in sheet_names:
                    cur.execute("INSERT INTO assignments (user_id, sheet_name, role) VALUES (%s, %s, %s)", (uid, sname, 'tham_tra'))
        wb.close()
        cur.execute("SELECT id FROM users WHERE username = %s", ('admin',))
        if not cur.fetchone():
            cur.execute("INSERT INTO users (username, password, fullname, role) VALUES (%s, %s, %s, %s)", ('admin', 'admin123', 'Quản trị viên', 'admin'))
        conn.commit()
    print("--- Đã nạp dữ liệu phân công thành công ---")

def create_or_get_user(conn, fullname, role):
    username = unsigned_user(fullname)
    cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE username = %s", (username,))
    user = cur.fetchone()
    if user:
        cur.execute("UPDATE users SET fullname = %s WHERE username = %s", (fullname, username))
        conn.commit()
        return user['id']
    cur.execute("INSERT INTO users (username, password, fullname, role) VALUES (%s, %s, %s, %s) RETURNING id", (username, '123', fullname, role))
    user_id = cur.fetchone()['id']
    conn.commit()
    return user_id

def is_evaluation_complete(sheet_name):
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) as cnt FROM evaluations WHERE sheet_name = %s AND col_letter = 'G' AND value != '' AND value IS NOT NULL", (sheet_name,))
        has_results = cur.fetchone()['cnt']
        cur.execute("SELECT reviewer_signature FROM suggestions WHERE sheet_name = %s", (sheet_name,))
        has_signature = cur.fetchone()
    sig_ok = (has_signature is not None and has_signature['reviewer_signature'] and has_signature['reviewer_signature'].strip() != '')
    return has_results > 0 and sig_ok

def ensure_archive_table():
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS archives (
                id SERIAL PRIMARY KEY,
                archive_date TEXT NOT NULL,
                table_name TEXT NOT NULL,
                row_data TEXT NOT NULL
            )
        """)
        conn.commit()

def archive_current_data():
    ensure_archive_table()
    archive_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM evaluations")
        for row in cur.fetchall():
            cur.execute("INSERT INTO archives (archive_date, table_name, row_data) VALUES (%s, %s, %s)", (archive_date, 'evaluations', json.dumps(dict(row), ensure_ascii=False)))
        cur.execute("SELECT * FROM review_comments")
        for row in cur.fetchall():
            cur.execute("INSERT INTO archives (archive_date, table_name, row_data) VALUES (%s, %s, %s)", (archive_date, 'review_comments', json.dumps(dict(row), ensure_ascii=False)))
        cur.execute("SELECT * FROM suggestions")
        for row in cur.fetchall():
            cur.execute("INSERT INTO archives (archive_date, table_name, row_data) VALUES (%s, %s, %s)", (archive_date, 'suggestions', json.dumps(dict(row), ensure_ascii=False)))
        conn.commit()
    print(f"Đã sao lưu dữ liệu vào archive ngày {archive_date}")

def reset_current_data():
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM evaluations")
        cur.execute("DELETE FROM review_comments")
        cur.execute("DELETE FROM suggestions")
        conn.commit()

# -------------------- ROUTES --------------------
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM users WHERE username = %s AND password = %s", (username, password))
            user = cur.fetchone()
        if user:
            session['user_id'] = user['id']
            session['fullname'] = user['fullname']
            session['role'] = user['role']
            return redirect(url_for('dashboard'))
        flash('Sai tài khoản hoặc mật khẩu (Mật khẩu mặc định: 123 cho user, admin: admin123)')
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT a.sheet_name, a.role,
                   COALESCE(s.locked_danh_gia, 0) as locked_danh_gia,
                   COALESCE(s.locked_tham_tra, 0) as locked_tham_tra
            FROM assignments a
            LEFT JOIN suggestions s ON a.sheet_name = s.sheet_name
            WHERE a.user_id = %s
        """, (session['user_id'],))
        assigns = cur.fetchall()
    eval_status = {}
    for ass in assigns:
        if ass['role'] == 'tham_tra':
            eval_status[ass['sheet_name']] = is_evaluation_complete(ass['sheet_name'])
    return render_template('dashboard.html', assignments=assigns, eval_status=eval_status)

@app.route('/form/<sheet_name>')
def evaluation_form(sheet_name):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT role FROM assignments WHERE user_id = %s AND sheet_name = %s", (session['user_id'], sheet_name))
        assign = cur.fetchone()
    if not assign:
        return "Bạn không có quyền truy cập biểu mẫu này", 403
    headers, rows, extra = get_sheet_data(sheet_name)
    if not headers:
        return f"Không tìm thấy sheet {sheet_name} trong file forms.xlsx", 404
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT row_index, col_letter, value FROM evaluations WHERE sheet_name = %s", (sheet_name,))
        db_rows = cur.fetchall()
        evals = {(r['row_index'], r['col_letter']): r['value'] for r in db_rows if r['row_index'] >= 10}
        saved_header = {(r['row_index'], r['col_letter']): r['value'] for r in db_rows if r['row_index'] < 10}
        cur.execute("SELECT row_index, comment FROM review_comments WHERE sheet_name = %s", (sheet_name,))
        comms = {r['row_index']: r['comment'] for r in cur.fetchall()}
        cur.execute("SELECT * FROM suggestions WHERE sheet_name = %s", (sheet_name,))
        s = cur.fetchone()
    suggestion = (s['suggestion'] if s else '') or ''
    reviewer_comment = (s['reviewer_comment'] if s else '') or ''
    reviewer_signature = (s['reviewer_signature'] if s else '') or ''
    checker_signature = (s['checker_signature'] if s else '') or ''
    locked_danh_gia = int(s['locked_danh_gia']) if s and s['locked_danh_gia'] is not None else 0
    locked_tham_tra = int(s['locked_tham_tra']) if s and s['locked_tham_tra'] is not None else 0
    return render_template('evaluation_form.html', sheet_name=sheet_name, role=assign['role'], headers=headers, rows=rows, extra=extra, saved=evals, saved_comments=comms, saved_header=saved_header, suggestion=suggestion, reviewer_comment=reviewer_comment, reviewer_signature=reviewer_signature, checker_signature=checker_signature, locked_danh_gia=locked_danh_gia, locked_tham_tra=locked_tham_tra, enumerate=enumerate)

@app.route('/evaluation/<sheet_name>')
def evaluation_redirect(sheet_name):
    return redirect(url_for('evaluation_form', sheet_name=sheet_name))

@app.route('/save', methods=['POST'])
def save():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    sn = request.form['sheet_name']
    role = request.form.get('role')
    uid = session['user_id']
    def col_name(col):
        return {'H': 'Mô tả', 'I': 'Đơn vị thực hiện', 'J': 'Thời gian', 'K': 'Giải pháp'}.get(col, col)
    with get_db_connection() as conn:
        cur = conn.cursor()
        if role == 'danh_gia':
            cycle_val = request.form.get('header_6_E', '').strip()
            if not cycle_val:
                flash('Vui lòng nhập "Nhiệt độ môi trường - Kiểm tra" (ô đầu tiên).')
                return redirect(url_for('evaluation_form', sheet_name=sn))
            cur.execute("""INSERT INTO evaluations (user_id, sheet_name, row_index, col_letter, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (user_id, sheet_name, row_index, col_letter) DO UPDATE SET value = EXCLUDED.value""", (uid, sn, 6, 'E', cycle_val))
        elif role == 'tham_tra':
            cycle_val = request.form.get('header_6_F', '').strip()
            if not cycle_val:
                flash('Vui lòng nhập "Nhiệt độ môi trường - Thẩm tra" (ô thứ hai).')
                return redirect(url_for('evaluation_form', sheet_name=sn))
            cur.execute("""INSERT INTO evaluations (user_id, sheet_name, row_index, col_letter, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (user_id, sheet_name, row_index, col_letter) DO UPDATE SET value = EXCLUDED.value""", (uid, sn, 6, 'F', cycle_val))
    eval_items = {}
    for key, value in request.form.items():
        if key.startswith('eval_'):
            parts = key.split('_')
            if len(parts) == 3:
                try:
                    row = int(parts[1])
                    col = parts[2]
                except:
                    continue
                if row not in eval_items:
                    eval_items[row] = {}
                eval_items[row][col] = value
    if role == 'danh_gia':
        for row, cols in eval_items.items():
            if 'G' not in cols or not cols['G'].strip():
                flash(f'Dòng {row}: chưa chọn kết quả (cột G).')
                return redirect(url_for('evaluation_form', sheet_name=sn))
            if cols.get('G') == 'K':
                missing = [col_name(c) for c in ['H','I','J','K'] if c not in cols or not cols[c].strip()]
                if missing:
                    flash(f'Dòng {row} (kết quả K) còn thiếu: {", ".join(missing)}.')
                    return redirect(url_for('evaluation_form', sheet_name=sn))
        reviewer_sig = request.form.get('reviewer_signature', '').strip()
        if not reviewer_sig:
            flash('Vui lòng nhập nội dung tại ô "Người đánh giá" (ký xác nhận).')
            return redirect(url_for('evaluation_form', sheet_name=sn))
        with get_db_connection() as conn:
            cur = conn.cursor()
            for k, v in request.form.items():
                if k.startswith('eval_'):
                    parts = k.split('_')
                    if len(parts) == 3:
                        try:
                            row = int(parts[1])
                            col = parts[2]
                        except:
                            continue
                        cur.execute("""INSERT INTO evaluations (user_id, sheet_name, row_index, col_letter, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (user_id, sheet_name, row_index, col_letter) DO UPDATE SET value = EXCLUDED.value""", (uid, sn, row, col, v))
            now = datetime.now().strftime('%Hh%M ngày %d/%m/%y')
            cur.execute("""INSERT INTO evaluations (user_id, sheet_name, row_index, col_letter, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (user_id, sheet_name, row_index, col_letter) DO UPDATE SET value = EXCLUDED.value""", (uid, sn, 4, 'F', now))
            cur.execute("""INSERT INTO suggestions (sheet_name, suggestion, reviewer_signature, locked_danh_gia) VALUES (%s, %s, %s, 1) ON CONFLICT (sheet_name) DO UPDATE SET suggestion = EXCLUDED.suggestion, reviewer_signature = EXCLUDED.reviewer_signature, locked_danh_gia = 1""", (sn, request.form.get('suggestion', ''), reviewer_sig))
        flash('Đã lưu đánh giá thành công.')
    elif role == 'tham_tra':
        comment_items = {}
        for key, value in request.form.items():
            if key.startswith('comment_'):
                parts = key.split('_')
                if len(parts) == 2:
                    try:
                        row = int(parts[1])
                    except:
                        continue
                    comment_items[row] = value
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT row_index FROM evaluations WHERE sheet_name = %s AND col_letter = 'G'", (sn,))
            rows_to_check = cur.fetchall()
        for r in rows_to_check:
            row = r['row_index']
            if row not in comment_items or not comment_items[row].strip():
                flash(f'Dòng {row}: chưa nhập ý kiến thẩm tra.')
                return redirect(url_for('evaluation_form', sheet_name=sn))
        checker_sig = request.form.get('checker_signature', '').strip()
        if not checker_sig:
            flash('Vui lòng nhập nội dung tại ô "Người thẩm tra" (ký xác nhận).')
            return redirect(url_for('evaluation_form', sheet_name=sn))
        with get_db_connection() as conn:
            cur = conn.cursor()
            for k, v in request.form.items():
                if k.startswith('comment_'):
                    parts = k.split('_')
                    if len(parts) == 2:
                        try:
                            row = int(parts[1])
                        except:
                            continue
                        cur.execute("""INSERT INTO review_comments (reviewer_id, sheet_name, row_index, comment) VALUES (%s, %s, %s, %s) ON CONFLICT (reviewer_id, sheet_name, row_index) DO UPDATE SET comment = EXCLUDED.comment""", (uid, sn, row, v))
            now = datetime.now().strftime('%Hh%M ngày %d/%m/%y')
            cur.execute("""INSERT INTO evaluations (user_id, sheet_name, row_index, col_letter, value) VALUES (%s, %s, %s, %s, %s) ON CONFLICT (user_id, sheet_name, row_index, col_letter) DO UPDATE SET value = EXCLUDED.value""", (uid, sn, 5, 'F', now))
            cur.execute("""INSERT INTO suggestions (sheet_name, reviewer_comment, checker_signature, locked_tham_tra) VALUES (%s, %s, %s, 1) ON CONFLICT (sheet_name) DO UPDATE SET reviewer_comment = EXCLUDED.reviewer_comment, checker_signature = EXCLUDED.checker_signature, locked_tham_tra = 1""", (sn, request.form.get('reviewer_comment', ''), checker_sig))
            cur.execute("SELECT row_index, col_letter, value FROM evaluations WHERE sheet_name = %s", (sn,))
            evals_snapshot = cur.fetchall()
            cur.execute("SELECT row_index, comment FROM review_comments WHERE sheet_name = %s", (sn,))
            comms_snapshot = cur.fetchall()
            cur.execute("SELECT * FROM suggestions WHERE sheet_name = %s", (sn,))
            sugg = cur.fetchone()
            snapshot = json.dumps({'evals': [{'row': r['row_index'], 'col': r['col_letter'], 'value': r['value']} for r in evals_snapshot], 'comments': [{'row': r['row_index'], 'comment': r['comment']} for r in comms_snapshot], 'suggestions': dict(sugg) if sugg else {}}, ensure_ascii=False)
            cur.execute("INSERT INTO history (sheet_name, role, user_id, snapshot) VALUES (%s, %s, %s, %s)", (sn, role, uid, snapshot))
        flash('Đã lưu ý kiến thẩm tra thành công.')
    return redirect(url_for('evaluation_form', sheet_name=sn))

# -------------------- Các route còn lại (giữ nguyên) --------------------
@app.route('/history')
def history():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    selected_month = request.args.get('month')
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT TO_CHAR(saved_at, 'YYYY-MM') as month FROM history ORDER BY month DESC")
        months = cur.fetchall()
        if selected_month:
            cur.execute("SELECT h.id, h.sheet_name, h.role, h.saved_at, u.fullname FROM history h JOIN users u ON h.user_id = u.id WHERE TO_CHAR(h.saved_at, 'YYYY-MM') = %s ORDER BY h.saved_at DESC", (selected_month,))
            rows = cur.fetchall()
        else:
            cur.execute("SELECT h.id, h.sheet_name, h.role, h.saved_at, u.fullname FROM history h JOIN users u ON h.user_id = u.id ORDER BY h.saved_at DESC")
            rows = cur.fetchall()
    return render_template('history.html', history=rows, months=months, selected_month=selected_month)

@app.route('/view_history/<int:history_id>')
def view_history(history_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT h.*, u.fullname as user_fullname FROM history h JOIN users u ON h.user_id = u.id WHERE h.id = %s", (history_id,))
        h = cur.fetchone()
        if not h:
            flash('Không tìm thấy bản ghi lịch sử.')
            return redirect(url_for('history'))
        snapshot = json.loads(h['snapshot'])
        headers, rows, extra = get_sheet_data(h['sheet_name'])
        return render_template('view_history.html', history=h, snapshot=snapshot, headers=headers, rows=rows, extra=extra, enumerate=enumerate)

@app.route('/export_all_forms')
def export_all_forms():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập chức năng này.')
        return redirect(url_for('dashboard'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT sheet_name FROM suggestions WHERE locked_tham_tra = 1")
        sheets = cur.fetchall()
    if not sheets:
        flash('Chưa có biểu mẫu nào được thẩm tra hoàn thành.')
        return redirect(url_for('dashboard'))
    wb = Workbook()
    wb.remove(wb.active)
    rev_map = build_reverse_mapping()
    for sheet in sheets:
        sheet_name = sheet['sheet_name']
        headers, rows, extra = get_sheet_data(sheet_name)
        if not headers:
            continue
        display_name = rev_map.get(sheet_name, sheet_name)
        ws = wb.create_sheet(title=display_name[:31])
        for i, row in enumerate(headers, start=1):
            ws.append([row.get(col, '') for col in 'ABCDEF'])
        ws.append([])
        ws.append(["Hạng mục", "STT", "Nội dung đánh giá", "Tiêu chuẩn", "Phương pháp", "Trạng thái TB", "Kết quả", "Mô tả", "Đơn vị thực hiện", "Thời gian", "Giải pháp"])
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute("SELECT row_index, col_letter, value FROM evaluations WHERE sheet_name = %s", (sheet_name,))
            evals = {(r['row_index'], r['col_letter']): r['value'] for r in cur.fetchall()}
            cur.execute("SELECT row_index, comment FROM review_comments WHERE sheet_name = %s", (sheet_name,))
            comments = {r['row_index']: r['comment'] for r in cur.fetchall()}
            cur.execute("SELECT * FROM suggestions WHERE sheet_name = %s", (sheet_name,))
            sugg = cur.fetchone()
        for idx, row in enumerate(rows, start=10):
            ws.append([row['A'], row['B'], row['C'], row['D'], row['E'], row['F'], evals.get((idx, 'G'), ''), evals.get((idx, 'H'), ''), evals.get((idx, 'I'), ''), evals.get((idx, 'J'), ''), evals.get((idx, 'K'), '')])
            if comments.get(idx):
                ws.cell(row=ws.max_row, column=12, value=comments[idx])
        ws.append([])
        ws.append(["Kiến nghị và ký xác nhận"])
        ws.append(["Kiến nghị (nếu có):", extra[0].get('B', '') if extra else ''])
        ws.append(["Người đánh giá:", sugg['reviewer_signature'] if sugg else ''])
        ws.append(["Người thẩm tra:", sugg['checker_signature'] if sugg else ''])
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'All_Forms_{datetime.now().strftime("%Y%m")}.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reset_cycle')
def reset_cycle():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập chức năng này.')
        return redirect(url_for('dashboard'))
    return '''
    <!DOCTYPE html><html><head><title>Xác nhận reset chu kỳ</title></head>
    <body>
        <h2>Bạn có chắc chắn muốn reset dữ liệu cho chu kỳ mới?</h2>
        <p>Dữ liệu hiện tại sẽ được sao lưu và xóa để bắt đầu chu kỳ đánh giá mới.</p>
        <form method="post" action="/confirm_reset">
            <button type="submit">Xác nhận reset</button>
            <a href="/dashboard">Hủy</a>
        </form>
    </body></html>
    '''

@app.route('/confirm_reset', methods=['POST'])
def confirm_reset():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập chức năng này.')
        return redirect(url_for('dashboard'))
    archive_current_data()
    reset_current_data()
    flash('Đã sao lưu và reset dữ liệu cho chu kỳ mới.')
    return redirect(url_for('dashboard'))

@app.route('/export_summary')
def export_summary():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập chức năng này.')
        return redirect(url_for('dashboard'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT e.sheet_name, e.row_index, e.value as result,
                   (SELECT value FROM evaluations e2 WHERE e2.sheet_name = e.sheet_name AND e2.row_index = e.row_index AND e2.col_letter = 'H') as description,
                   (SELECT comment FROM review_comments rc WHERE rc.sheet_name = e.sheet_name AND rc.row_index = e.row_index) as reviewer_comment
            FROM evaluations e WHERE e.col_letter = 'G' AND e.value = 'K' ORDER BY e.sheet_name, e.row_index
        """)
        rows = cur.fetchall()
        cur.execute("""
            SELECT sheet_name, reviewer_signature, checker_signature FROM suggestions
            WHERE sheet_name IN (SELECT DISTINCT sheet_name FROM evaluations WHERE col_letter = 'G' AND value = 'K')
        """)
        suggestions = cur.fetchall()
        sug_dict = {s['sheet_name']: (s['reviewer_signature'], s['checker_signature']) for s in suggestions}
    if not rows:
        flash('Không có khiếm khuyết nào để xuất báo cáo.')
        return redirect(url_for('dashboard'))
    rev_map = build_reverse_mapping()
    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng hợp KKTB"
    ws.merge_cells('A1:D1')
    ws['A1'] = "TỔNG HỢP KHIẾM KHUYẾT THIẾT BỊ TPM"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append(["STT", "Biểu mẫu", "Nội dung khiếm khuyết (Mô tả của ĐG)", "Mô tả của Thẩm tra"])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for stt, row in enumerate(rows, start=1):
        ws.append([stt, rev_map.get(row['sheet_name'], row['sheet_name']), row['description'] or '', row['reviewer_comment'] or ''])
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        if col[0].column_letter:
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    if sug_dict:
        ws.append([])
        ws.append(["KIẾN NGHỊ VÀ Ý KIẾN THẨM TRA"])
        ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal='center')
        for sc, (rs, cc) in sug_dict.items():
            ws.append([sc, "Kiến nghị của người đánh giá:", rs or '', ""])
            ws.append([sc, "Ý kiến của người thẩm tra:", cc or '', ""])
            for r in range(ws.max_row - 1, ws.max_row + 1):
                for cell in ws[r]:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='TonghopKKTB TPM.xlsx', as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập trang này.')
        return redirect(url_for('dashboard'))
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT sheet_name FROM evaluations WHERE col_letter = 'G' AND value != ''")
        sheets = cur.fetchall()
        system_data = []
        for sheet in sheets:
            sn = sheet['sheet_name']
            cur.execute("SELECT COUNT(*) as cnt FROM evaluations WHERE sheet_name = %s AND col_letter = 'G' AND value = 'K'", (sn,))
            k_count = cur.fetchone()['cnt']
            cur.execute("SELECT COUNT(*) as cnt FROM evaluations WHERE sheet_name = %s AND col_letter = 'G' AND value != ''", (sn,))
            total = cur.fetchone()['cnt']
            system_data.append({'name': sn, 'total': total, 'k_count': k_count, 'percentage': round((k_count / total * 100) if total > 0 else 0, 1)})
        system_data.sort(key=lambda x: x['percentage'], reverse=True)
    return render_template('admin_dashboard.html', system_data=system_data)

@app.route('/sync_assignments', methods=['POST'])
def sync_assignments():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Bạn không có quyền truy cập chức năng này.')
        return redirect(url_for('dashboard'))
    if not os.path.exists(PHAN_GIAO_FILE):
        flash('Không tìm thấy file phan_giao.xlsx')
        return redirect(url_for('dashboard'))
    mapping = build_sheet_mapping()
    new_assignments = []
    wb = safe_load_workbook(PHAN_GIAO_FILE)
    if not wb:
        flash('Không thể đọc file phan_giao.xlsx')
        return redirect(url_for('dashboard'))
    ws = wb.active
    for row in range(8, ws.max_row + 1):
        ma_bieu_mau = str(ws[f'C{row}'].value or '').strip()
        name_eval = str(ws[f'E{row}'].value or '').strip()
        name_check = str(ws[f'F{row}'].value or '').strip()
        if not ma_bieu_mau or (not name_eval and not name_check):
            continue
        if ma_bieu_mau in ['BM.P4.15.18', 'BM.P4.15.19']:
            base_num = ma_bieu_mau.split('.')[-1]
            snames = [f'BM{base_num}_a', f'BM{base_num}_b', f'BM{base_num}_c']
        else:
            sname = mapping.get(ma_bieu_mau)
            if not sname:
                continue
            snames = [sname]
        for sn in snames:
            if name_eval:
                new_assignments.append((name_eval, sn, 'danh_gia'))
            if name_check:
                new_assignments.append((name_check, sn, 'tham_tra'))
    wb.close()
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute("SELECT a.id, a.sheet_name, a.role, u.fullname FROM assignments a JOIN users u ON a.user_id = u.id")
        current = cur.fetchall()
        current_set = {(r['fullname'], r['sheet_name'], r['role']) for r in current}
        new_set = set(new_assignments)
        added, removed = [], []
        for (fullname, sn, role) in new_set - current_set:
            uid = create_or_get_user(conn, fullname, role)
            cur.execute("INSERT INTO assignments (user_id, sheet_name, role) VALUES (%s, %s, %s)", (uid, sn, role))
            added.append(f"{fullname} - {sn} ({'đánh giá' if role=='danh_gia' else 'thẩm tra'})")
        for r in current:
            if (r['fullname'], r['sheet_name'], r['role']) not in new_set:
                cur.execute("DELETE FROM assignments WHERE id = %s", (r['id'],))
                removed.append(f"{r['fullname']} - {r['sheet_name']} ({'đánh giá' if r['role']=='danh_gia' else 'thẩm tra'})")
    flash('Đã đồng bộ phân công từ file phan_giao.xlsx.')
    if added:
        flash(f'✅ Thêm mới {len(added)} bản ghi:')
        for item in added[:20]:
            flash(f'  + {item}')
        if len(added) > 20:
            flash(f'  ... và {len(added)-20} bản ghi khác')
    if removed:
        flash(f'❌ Xóa bỏ {len(removed)} bản ghi:')
        for item in removed[:20]:
            flash(f'  - {item}')
        if len(removed) > 20:
            flash(f'  ... và {len(removed)-20} bản ghi khác')
    if not added and not removed:
        flash('Không có thay đổi nào so với phân công hiện tại.')
    return redirect(url_for('dashboard'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    # Chạy lần đầu để tạo bảng: bỏ comment dòng dưới, chạy xong comment lại
    # init_db()
    app.run(debug=True, host='0.0.0.0')